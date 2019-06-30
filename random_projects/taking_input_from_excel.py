import openpyxl
from simple_salesforce import Salesforce

data_list = []

def create_mass_contacts():

	for i in data_list:

		temp_last_name = i["lastname"]
		temp_email = i["email"]
		sf.Contact.create({"lastname":temp_last_name, "email":temp_email})

wb = openpyxl.load_workbook("contact_list_file_1.xlsx")

sheet1= wb.active

for row in sheet1.iter_rows(min_row = 2, max_col = 3, max_row = sheet1.max_row, values_only = True):
	data_list.append({"lastname":str(row[1]),"email":str(row[2])})

print(data_list)

sf = Salesforce(instance_url='https://test.salesforce.com', session_id='')

sf_username = "mbit@apparo.org.refreshbox"
sf_password = "Amine04marwan"
sf_org_id = "00D3F000000DP88"
sf_security_token = "nqVE3uOrLVrFdvDXIIMd0LsbE"

sf = Salesforce(username = sf_username, password = sf_password, organizationId = sf_org_id, security_token = sf_security_token, domain = "test" )

create_mass_contacts()