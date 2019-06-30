
import openpyxl

wb = openpyxl.load_workbook("contact_list_file.xlsx")

print(wb.get_sheet_names())
print(wb.active)

sheet = wb.active 

print(sheet["A1"].value)
print(sheet["B1"].value)
print(sheet["C1"].value)

contact = sheet["A1"]
print("(", contact.row, ",", contact.column, ")")

print(sheet.cell(row=2, column=3))
print(sheet.cell(row=2, column=3).value)
